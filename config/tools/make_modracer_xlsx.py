# -*- coding: utf-8 -*-
"""
ModRacer 配表生成脚本（仅首次建表用）。

生成 config/data/ModRacer.xlsx，包含 3 张表：
  - 车型-car    基础载具底盘
  - 改件-part   性能/战术改件
  - 地图-map    赛道与词缀

表结构遵循 GDExcelExporter 规范：每 sheet 三行表头
  第1行：字段类型（int/float/string/bool/tr_string...）
  第2行：中文注释（导出为代码注释）
  第3行：字段名（首列必须为 id；* 前缀字段不导出）
之后数值迭代请直接在 Excel 中改表，不要再跑本脚本（会覆盖）。
"""
from openpyxl import Workbook

wb = Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------- 车型表
car = wb.create_sheet("车型-car")
car.append(["int", "tr_string", "string", "int", "float", "float", "int", "int", "int", "tr_string"])
car.append(["编号", "车型名", "驱动形式", "极速km/h", "加速(0-10)", "操控(0-10)", "整备质量kg", "性能槽数", "功能槽数", "定位描述"])
car.append(["id", "name", "drive", "top_speed", "accel", "handling", "weight", "perf_slots", "func_slots", "desc"])
car_rows = [
    [1, "狂暴马力型", "RWD", 320, 7.5, 5.5, 1500, 4, 1, "长直道霸主，撞击抗性好；低速易打滑，雨雪泥地操控难。"],
    [2, "灵巧敏捷型", "FWD", 260, 7.0, 9.0, 1100, 4, 1, "多弯赛道极其稳定，容错率高；极速上限低，直线弱。"],
    [3, "全能均衡型", "AWD", 290, 8.0, 7.5, 1300, 3, 2, "起步抓地强，全地形适应好；无明显极端长板。"],
]
for row in car_rows:
    car.append(row)

# ---------------------------------------------------------------- 改件表
part = wb.create_sheet("改件-part")
part.append([
    "int", "tr_string", "string", "int",
    "float", "float", "float", "float", "float", "float", "float", "int",
    "float", "int", "float", "tr_string",
])
part.append([
    "编号", "改件名", "类别", "稀有度1-4",
    "极速加成%", "加速加成%", "铺装抓地%", "越野抓地%", "雨雪防滑%", "气动下压%", "落地稳定%", "质量增减kg",
    "技能CD秒", "单回合弹药", "效果持续秒", "效果描述",
])
part.append([
    "id", "name", "category", "rarity",
    "top_speed", "accel", "grip_road", "grip_offroad", "grip_wet", "aero", "landing", "mass",
    "cooldown", "ammo", "duration", "desc",
])
part_rows = [
    # 引擎
    [101, "自然吸气强化引擎", "engine", 1, 3, 3, 0, 0, 0, 0, 0, 0, 0, 0, 0, "均衡小幅提升动力。"],
    [102, "机械增压引擎", "engine", 2, 6, 5, 0, 0, 0, 0, 0, 10, 0, 0, 0, "中段加速凶悍，略增重量。"],
    [103, "大马力涡轮引擎", "engine", 3, 10, 7, 0, -2, 0, 0, 0, 20, 0, 0, 0, "极限极速取向，低速扭矩迟滞。"],
    # 轮胎
    [201, "运动热熔胎", "tires", 2, 0, 2, 8, -4, -2, 0, 0, 0, 0, 0, 0, "铺装路抓地极强，越野雨雪衰减。"],
    [202, "拉力越野胎", "tires", 2, -1, 0, -2, 12, 2, 0, 0, 0, 0, 0, 0, "砂石泥地表现优异，极速略降。"],
    [203, "深纹雨胎", "tires", 2, -1, 0, -1, 2, 12, 0, 0, 0, 0, 0, 0, "暴雨湿滑路面防滑利器。"],
    [204, "全地形复合胎", "tires", 3, 0, 1, 4, 6, 6, 0, 0, 0, 0, 0, 0, "无明显短板的全能胎。"],
    # 尾翼/气动
    [301, "低阻尾翼", "aero", 1, 4, 0, 0, 0, 0, -3, 0, 0, 0, 0, 0, "直线减阻，冲刺极速提升。"],
    [302, "大下压力尾翼", "aero", 2, -2, 1, 3, 1, 1, 8, 0, 0, 0, 0, 0, "高速过弯稳定性大幅提升。"],
    # 悬挂/底盘
    [401, "强化悬挂", "chassis", 2, 0, 0, 1, 3, 0, 0, 10, 15, 0, 0, 0, "飞跳落地稳定，抗撞击。"],
    [402, "轻量化底盘", "chassis", 3, 2, 3, 0, 0, 0, 0, 3, -80, 0, 0, 0, "减重提速能，抗撞变弱。"],
    # 战术/功能
    [501, "火箭筒", "tactical", 2, 0, 0, 0, 0, 0, 0, 0, 0, 20, 2, 0, "锁定前方最近车辆发射，造成减速打转。"],
    [502, "战术隐身", "tactical", 3, 0, 0, 0, 0, 0, 0, 0, 0, 30, 1, 5, "半透明且无法被锁定，免疫追踪攻击。"],
    [503, "超级氮气", "tactical", 2, 0, 0, 0, 0, 0, 0, 0, 0, 15, 3, 3, "瞬间喷射提供极强推力。"],
]
for row in part_rows:
    part.append(row)

# ---------------------------------------------------------------- 地图表
map_ = wb.create_sheet("地图-map")
map_.append(["int", "tr_string", "string", "string", "int", "int", "int", "bool", "tr_string"])
map_.append(["编号", "地图名", "地形", "天气词缀", "直道占比%", "弯道数", "跳台数", "有高风险分支", "地图描述"])
map_.append(["id", "name", "terrain", "weather", "straight_ratio", "corner_count", "jump_count", "hazard_branch", "desc"])
map_rows = [
    [1, "环湖高速公路", "asphalt", "sunny", 70, 4, 0, False, "长直道为主，极速车的天堂。"],
    [2, "砂石荒漠峡谷", "gravel", "sandstorm", 40, 8, 6, True, "陡坡跳台+泥地打滑，越野胎与强化悬挂主场。"],
    [3, "雨雾山道", "asphalt", "storm", 30, 14, 2, True, "暴雨湿滑连续弯，雨胎+大下压尾翼克制。"],
    [4, "冰雪极地走廊", "snow", "snow", 50, 9, 4, True, "低温积雪路面，抓地与防滑是生死线。"],
]
for row in map_rows:
    map_.append(row)

# 首页说明（~ 开头不会被导出）
cover = wb.create_sheet("~首页", 0)
cover["B2"] = "ModRacer 配表"
cover["B3"] = "本文件为游戏数值唯一真理源头，修改后运行 ee gen-all 导出。"
cover["B4"] = "表结构规范见 config/tools/make_modracer_xlsx.py 头部注释。"

wb.save("data/ModRacer.xlsx")
print("已生成 data/ModRacer.xlsx")
