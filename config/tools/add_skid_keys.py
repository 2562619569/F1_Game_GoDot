# -*- coding: utf-8 -*-
"""
ModRacer 配表升级脚本（2026-08）：极限工况车轮印参数（SkidMarks skid_*）。

  - Game-game 补 6 行 skid 键（id 动态分配，接在现有最大 id 后——
    并行功能也会加键，硬编码 id 段会撞号，见 env_* 抢占 35~44 的教训）：
      skid_lat_slip    侧滑角阈值（rad ≈11.5°：漂移/甩尾横向滑动超过才落印）
      skid_lon_slip    纵滑率阈值（刹车抱死为正、烧胎打滑为大负）
      skid_lifetime    车轮印淡出时长（s）
      skid_alpha       满强度不透明度
      skid_gap         最小分段长（m，触点移动超过才铺新段）
      skid_pool        环形缓冲段数（四轮共享，用尽覆写最旧）
按 key 判重幂等。用法:  cd config && python tools/add_skid_keys.py
"""
import openpyxl

PATH = "data/ModRacer.xlsx"
ADD_ROWS = [
    ("skid_lat_slip", 0.20, "Lateral slip angle threshold for tire marks (rad, ~11.5deg)"),
    ("skid_lon_slip", 0.20, "Longitudinal slip threshold for tire marks (lockup/wheelspin)"),
    ("skid_lifetime", 25.0, "Tire mark fade-out duration (sec)"),
    ("skid_alpha", 0.75, "Tire mark peak opacity"),
    ("skid_gap", 0.35, "Min segment length between mark quads (m)"),
    ("skid_pool", 4096.0, "Ring buffer segments per car (oldest overwritten)"),
]

wb = openpyxl.load_workbook(PATH)
ws = wb["Game-game"]
assert ws.cell(row=3, column=1).value == "id" and ws.cell(row=3, column=2).value == "key", \
    "Game 表结构与预期不符，中止"

# 清掉数据区内的占位空行：id 列为空/空白串/0 即删（样式残留行没有值但占行，
# 按全列判 None 清不干净，会被导出器当成 id=0 空记录——2026-08 实测两度复现）
r = 4
while r <= ws.max_row:
    v = ws.cell(row=r, column=1).value
    if v is None or str(v).strip() == "" or str(v).strip() == "0":
        ws.delete_rows(r)
    else:
        r += 1

# 撞号自愈：并行功能各自加键会硬编码同段 id（env_* 与 skid_* 曾同时占 35 段），
# 导出字典按 id 后行覆盖前行，先到者会被静默吞掉——扫一遍把重复 id 重排到
# 现有最大 id 之后，行序不变
seen = {}
next_id = 0
row = 4
while ws.cell(row=row, column=1).value is not None:
    rid = int(ws.cell(row=row, column=1).value)
    next_id = max(next_id, rid)
    if rid in seen:
        next_id += 1
        print("id %d 撞号（%s），重排为 %d" % (rid, ws.cell(row=row, column=2).value, next_id))
        ws.cell(row=row, column=1).value = next_id
    else:
        seen[rid] = row
    row += 1

keys = {}
row = 4
while ws.cell(row=row, column=1).value is not None:
    keys[ws.cell(row=row, column=2).value] = row
    next_id = max(next_id, int(ws.cell(row=row, column=1).value))
    row += 1

added = 0
for key, value, note in ADD_ROWS:
    if key in keys:
        print("Game-game 已有 %s，跳过" % key)
        continue
    next_id += 1
    ws.append([next_id, key, value, note])
    added += 1

wb.save(PATH)
print("Game-game 新增 %d 行（skid_*，id %d 起），总行数 %d"
      % (added, next_id - added + 1, row - 4 + added))
