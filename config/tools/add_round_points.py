# -*- coding: utf-8 -*-
"""
ModRacer 配表升级脚本（2026-08）：回合积分制（积分总冠军取代决赛一局定胜负）。

  - Round-round 表加 points 列（array，名次 -> 本回合积分，`|` 分隔，索引 = 名次-1）：
      名次超出数组长度按 0 分处理；常规回合小分差，决赛行配高额前倾分布制造翻盘空间。
  - 数值依据（4 车制，可随时改表调参）：
      常规 3|2|1|0       ：三回合最多拉开 9 分
      决赛 12|8|4|0      ：P1-P4 分差 12 > 9（全垫底者决赛夺冠即翻盘）
                           P1-P2 分差 4 > 3（三连亚 vs 三连胜，决赛反超 1 分）
用法:  cd config && python tools/add_round_points.py
"""
import openpyxl

PATH = "data/ModRacer.xlsx"

# 每回合积分表：回合 id -> `|` 分隔的积分串（补齐到满编 8 名，5~8 名 0 分）
POINTS = {
    1: "3|2|1|0|0|0|0|0",
    2: "3|2|1|0|0|0|0|0",
    3: "3|2|1|0|0|0|0|0",
    4: "12|8|4|0|0|0|0|0",
}

wb = openpyxl.load_workbook(PATH)
ws = wb["Round-round"]
assert ws.cell(row=3, column=1).value == "id" and ws.cell(row=3, column=5).value == "map_pool", \
    "Round 表结构与预期不符，中止"

if "points" in [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]:
    print("Round-round 已有 points 列，跳过")
    raise SystemExit(0)

col = ws.max_column + 1
ws.cell(row=1, column=col).value = "array"
ws.cell(row=2, column=col).value = "名次积分表(名次1起,超长按0,决赛加码供翻盘)"
ws.cell(row=3, column=col).value = "points"

row = 4
touched = 0
while ws.cell(row=row, column=1).value is not None:
    rid = int(ws.cell(row=row, column=1).value)
    if rid in POINTS:
        ws.cell(row=row, column=col).value = POINTS[rid]
        touched += 1
    row += 1

wb.save(PATH)
print("Round-round 第 %d 列新增 points（写入 %d 行），下一可用 id 行=%d" % (col, touched, row))
