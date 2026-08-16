# -*- coding: utf-8 -*-
"""
ModRacer 配表重构脚本（一次性，2026-08）。

规范：值与字段名只允许英文；第 2 行为中文注释行，不会被导出器读取。
  1. 全部单元格值英文化
  2. Map 表精简为 id/name/desc（地图名/介绍）；天气/环境已迁移到地图 env 文件
     （game/race/tracks/data/map_<id>_env.json，见 WeatherEnv）
  3. Car 表按 addons/gevp/scripts/vehicle.gd 的导出参数扩充物理基础参数：
     max_torque, max_rpm, final_drive, gear_ratios, front_torque_split,
     max_steering_angle, steering_speed, brake_force_multiplier,
     coefficient_of_drag, frontal_area, front_weight_distribution,
     center_of_gravity_height_offset, inertia_multiplier
可重复执行（幂等，整簿重建）。
⚠️ 注意：本脚本数据是历史快照，整簿重建会覆盖 xlsx 后续演进（曾把 Car 601 段重建回 1/2/3 造成线上报错）。
   xlsx 才是唯一真理源头，日常改动请直接改 xlsx 或写手术式脚本（参考 add_cosmetic_wheel.py）。
"""
from openpyxl import Workbook

PATH = "data/ModRacer.xlsx"

wb = Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------- Car
car = wb.create_sheet("Car-car")
car.append([
    "int", "tr_string", "string", "int", "float", "float", "int", "int", "int", "tr_string", "int", "int",
    "float", "float", "float", "array", "float",
    "float", "float", "float", "float", "float", "float", "float", "float",
])
car.append([
    "编号", "车型名", "驱动形式", "极速km/h", "加速(0-10)", "操控(0-10)",
    "整备质量kg", "性能槽数", "功能槽数", "定位描述", "铺装抓地(0-10)", "越野抗性(0-10)",
    "最大扭矩NM", "最大转速rpm", "主减速比", "各挡齿比(|分隔)", "前轴扭矩占比(0=RWD 1=FWD)",
    "最大转向角(度)", "转向速率", "制动力倍率", "风阻系数", "迎风面积m2", "前轴配重(0-1)",
    "重心高度偏移m", "转动惯量倍率",
])
car.append([
    "id", "name", "drive", "top_speed", "accel", "handling", "weight", "perf_slots", "func_slots", "desc",
    "grip_road", "grip_offroad",
    "max_torque", "max_rpm", "final_drive", "gear_ratios", "front_torque_split",
    "max_steering_angle", "steering_speed", "brake_force_multiplier",
    "coefficient_of_drag", "frontal_area", "front_weight_distribution",
    "center_of_gravity_height_offset", "inertia_multiplier",
])
car_rows = [
    [601, "Brute Power", "RWD", 320, 7.5, 5.5, 1500, 4, 1,
     "Straight-line monster with great impact resistance; slippery at low speed, tricky on wet/snow/mud.",
     6, 3, 420.0, 7500.0, 3.4, "3.6|2.2|1.6|1.25|1.0|0.8", 0.0,
     36.0, 3.5, 1.0, 0.32, 2.1, 0.48, -0.15, 1.3],
    [602, "Agile Sprinter", "FWD", 260, 7.0, 9.0, 1100, 4, 1,
     "Rock solid on twisty tracks with high forgiveness; low top speed, weak on straights.",
     8, 5, 260.0, 6800.0, 3.6, "3.8|2.4|1.7|1.3|1.0", 1.0,
     44.0, 5.5, 1.1, 0.34, 1.9, 0.62, -0.25, 1.0],
    [603, "All-Rounder", "AWD", 290, 8.0, 7.5, 1300, 3, 2,
     "Strong launch grip and all-terrain adaptability; no extreme strengths.",
     7, 8, 330.0, 7000.0, 3.5, "3.7|2.3|1.65|1.28|1.0|0.82", 0.5,
     40.0, 4.5, 1.05, 0.33, 2.0, 0.52, -0.20, 1.15],
]
for row in car_rows:
    car.append(row)

# ---------------------------------------------------------------- Part
part = wb.create_sheet("Part-part")
part.append([
    "int", "tr_string", "string", "int",
    "float", "float", "float", "float", "float", "float", "float", "int",
    "float", "int", "float", "tr_string", "string", "float",
])
part.append([
    "编号", "改件名", "类别", "稀有度1-4",
    "极速加成%", "加速加成%", "铺装抓地%", "越野抓地%", "雨雪防滑%", "气动下压%", "落地稳定%", "质量增减kg",
    "技能CD秒", "单回合弹药", "效果持续秒", "效果描述", "效果枚举", "效果强度(按effect解释)",
])
part.append([
    "id", "name", "category", "rarity",
    "top_speed", "accel", "grip_road", "grip_offroad", "grip_wet", "aero", "landing", "mass",
    "cooldown", "ammo", "duration", "desc", "effect", "power",
])
part_rows = [
    # engine
    [101, "NA Boost Engine", "engine", 1, 3, 3, 0, 0, 0, 0, 0, 0, 0, 0, 0, "Balanced mild power gain.", "none", 0.0],
    [102, "Supercharged Engine", "engine", 2, 6, 5, 0, 0, 0, 0, 0, 10, 0, 0, 0, "Fierce mid-range pull, slightly heavier.", "none", 0.0],
    [103, "Big Turbo Engine", "engine", 3, 10, 7, 0, -2, 0, 0, 0, 20, 0, 0, 0, "Top-speed oriented, low-end lag.", "none", 0.0],
    # tires
    [201, "Sport Slicks", "tires", 2, 0, 2, 8, -4, -2, 0, 0, 0, 0, 0, 0, "Huge asphalt grip, weak offroad/wet.", "none", 0.0],
    [202, "Rally Offroad Tires", "tires", 2, -1, 0, -2, 12, 2, 0, 0, 0, 0, 0, 0, "Great on gravel/mud, slightly lower top speed.", "none", 0.0],
    [203, "Deep-Tread Rain Tires", "tires", 2, -1, 0, -1, 2, 12, 0, 0, 0, 0, 0, 0, "Anti-slip on wet roads.", "none", 0.0],
    [204, "All-Terrain Tires", "tires", 3, 0, 1, 4, 6, 6, 0, 0, 0, 0, 0, 0, "No weak spot all-rounder.", "none", 0.0],
    # aero
    [301, "Low-Drag Wing", "aero", 1, 4, 0, 0, 0, 0, -3, 0, 0, 0, 0, 0, "Less drag, higher sprint speed.", "none", 0.0],
    [302, "High-Downforce Wing", "aero", 2, -2, 1, 3, 1, 1, 8, 0, 0, 0, 0, 0, "Much more stable in high-speed corners.", "none", 0.0],
    # chassis
    [401, "Reinforced Suspension", "chassis", 2, 0, 0, 1, 3, 0, 0, 10, 15, 0, 0, 0, "Stable landings, impact resistant.", "none", 0.0],
    [402, "Lightweight Chassis", "chassis", 3, 2, 3, 0, 0, 0, 0, 3, -80, 0, 0, 0, "Less weight, quicker, weaker to impacts.", "none", 0.0],
    # tactical
    [501, "Rocket Launcher", "tactical", 2, 0, 0, 0, 0, 0, 0, 0, 0, 20, 2, 0, "Locks nearest car ahead, slows and spins it.", "slow_spin", 30.0],
    [502, "Tactical Stealth", "tactical", 3, 0, 0, 0, 0, 0, 0, 0, 0, 30, 1, 5, "Semi-transparent, untargetable, immune to tracking.", "stealth", 0.5],
    [503, "Super Nitrous", "tactical", 2, 0, 0, 0, 0, 0, 0, 0, 0, 15, 3, 3, "Instant strong thrust burst.", "nitro_push", 2.0],
]
for row in part_rows:
    part.append(row)

# ---------------------------------------------------------------- Map (simplified: name / desc; env moved to map env json)
map_ = wb.create_sheet("Map-map")
map_.append(["int", "tr_string", "tr_string"])
map_.append(["编号", "地图名", "地图介绍"])
map_.append(["id", "name", "desc"])
map_rows = [
    [1, "Lakeside Highway", "Long straights under open sky, top-speed heaven."],
    [2, "Desert Gravel Canyon", "Steep jumps and slippery dirt. Offroad tires and tough suspension shine."],
    [3, "Rainy Mountain Pass", "Storm-soaked tight corners. Rain tires and downforce prevail."],
    [4, "Frozen Polar Corridor", "Packed snow and ice; grip is everything."],
]
for row in map_rows:
    map_.append(row)

# ---------------------------------------------------------------- Game
game = wb.create_sheet("Game-game")
game.append(["int", "string", "float", "string"])
game.append(["编号", "参数名", "参数值", "说明"])
game.append(["id", "key", "value", "note"])
game_rows = [
    [1, "player_max", 8, "Max players (incl. AI)"],
    [2, "intermission_sec", 40, "Intermission duration (sec)"],
    [3, "round_count", 4, "Sub-round count"],
    [4, "start_countdown", 3, "Start countdown (sec)"],
    [5, "lock_ahead_range", 60, "Rocket lock-on range ahead (m)"],
    [6, "loot_pick_radius", 3, "Loot pickup radius (m)"],
]
for row in game_rows:
    game.append(row)

# ---------------------------------------------------------------- Round
round_ = wb.create_sheet("Round-round")
round_.append(["int", "tr_string", "bool", "int", "string"])
round_.append(["回合号", "回合名", "是否决战局", "时限秒", "候选地图id"])
round_.append(["id", "name", "is_final", "time_limit", "map_pool"])
round_rows = [
    [1, "Round 1", False, 240, "1|2"],
    [2, "Round 2", False, 270, "2|3"],
    [3, "Round 3", False, 270, "3|4"],
    [4, "Final Showdown", True, 300, "1|2|3|4"],
]
for row in round_rows:
    round_.append(row)

# ---------------------------------------------------------------- Loot
loot = wb.create_sheet("Loot-loot")
loot.append(["int", "string", "int", "string", "string", "int"])
loot.append(["编号", "路线类型", "掉落点数", "稀有度权重(r1|r2|r3|r4)", "可掉类别", "保底稀有度(0无)"])
loot.append(["id", "route", "drop_count", "rarity_weights", "category_pool", "guarantee_rarity"])
loot_rows = [
    [1, "main", 6, "60|30|10|0", "engine|tires|aero|chassis|tactical", 0],
    [2, "hazard", 3, "0|20|50|30", "engine|tires|aero|chassis|tactical", 2],
]
for row in loot_rows:
    loot.append(row)

# ---------------------------------------------------------------- RankReward
rr = wb.create_sheet("RankReward-rank_reward")
rr.append(["int", "int", "int", "int"])
rr.append(["名次", "奖励件数", "奖励最低稀有度", "下回合发车位"])
rr.append(["id", "reward_count", "reward_rarity_min", "grid_next"])
rr_rows = [
    [1, 2, 3, 8],
    [2, 2, 2, 7],
    [3, 1, 2, 6],
    [4, 1, 2, 5],
    [5, 1, 1, 4],
    [6, 1, 1, 3],
    [7, 1, 1, 2],
    [8, 1, 1, 1],
]
for row in rr_rows:
    rr.append(row)

# ---------------------------------------------------------------- Cover (ignored by exporter)
cover = wb.create_sheet("~Cover", 0)
cover["B2"] = "ModRacer Config Tables"
cover["B3"] = "Single source of truth for game values. Run 'ee gen-all' after edits."
cover["B4"] = "Values/field names in English only; row 2 comments may be Chinese (not exported)."

wb.save(PATH)
print("Rebuilt", PATH, ":", ", ".join(wb.sheetnames))
