# -*- coding: utf-8 -*-
"""
ModRacer 配表升级脚本（2026-08）：体积云天空参数（VolumetricClouds）。

  - Game-game 补 4 行云键（紧接当时表尾 id，按 key 判重幂等）：
      env_clouds_enabled   体积云天空总开关（1=clayjohn 移植 raymarch 天空）
      env_cloud_coverage   云量基础值 0.1~1（各天气预设再乘修正：风暴厚晴天疏）
      env_cloud_density    云密度（消光系数基础值）
      env_cloud_wind       云风速（视运动快慢）
用法:  cd config && python tools/add_cloud_keys.py
"""
import openpyxl

PATH = "data/ModRacer.xlsx"
ADD_KEYS = [
    ("env_clouds_enabled", 1.0, "Volumetric cloud sky master switch (clayjohn raymarch sky shader)"),
    ("env_cloud_coverage", 0.30, "Base cloud coverage 0.1-1 (per-weather preset multiplies on top)"),
    ("env_cloud_density", 0.055, "Base cloud density (absorption)"),
    ("env_cloud_wind", 2.5, "Cloud wind speed"),
]

wb = openpyxl.load_workbook(PATH)
ws = wb["Game-game"]
assert ws.cell(row=3, column=1).value == "id" and ws.cell(row=3, column=2).value == "key", \
    "Game 表结构与预期不符，中止"

# 清掉数据区中间/尾部的全空行（历史格式残留会被导出器当成 id=0 空记录）
r = 4
while r <= ws.max_row:
    if all(ws.cell(row=r, column=c).value is None for c in range(1, ws.max_column + 1)):
        ws.delete_rows(r)
    else:
        r += 1

keys = {}
max_id = 0
row = 4
while ws.cell(row=row, column=1).value is not None:
    keys[ws.cell(row=row, column=2).value] = row
    max_id = max(max_id, int(ws.cell(row=row, column=1).value))
    row += 1

added = 0
for key, value, note in ADD_KEYS:
    if key in keys:
        print("Game-game 已有 %s，跳过" % key)
        continue
    max_id += 1
    ws.append([max_id, key, value, note])
    added += 1

wb.save(PATH)
print("Game-game 新增 %d 行（env_cloud_*），下一可用 id=%d" % (added, max_id + 1))
