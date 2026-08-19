# -*- coding: utf-8 -*-
"""
ModRacer 配表升级脚本（2026-08）：空格漂移模式参数（DriftMode drift_*）。

  - Game-game 补 6 行 drift 键（id 29~34）：
      drift_speed_min    进入/维持漂移的最低车速 m/s（发车锁车速度≈0 天然防误触）
      drift_brake_scale  漂移中手刹制动力缩放（起漂咬一下而非锁轮急刹）
      drift_rear_grip    漂移中后轮侧向抓地缩放（前轮不动保指向）
      drift_slip_assist  漂移中转向滑移辅助阈值 rad（放宽大角度打方向）
      drift_yaw_engage   漂移中横摆稳定介入角（dot 域，≈45° 车速偏差才纠偏兜底）
      drift_yaw_kick     起漂甩尾角速度增量 rad/s（按转向输入取方向与比例）
按 key 判重幂等。用法:  cd config && python tools/add_drift_keys.py
"""
import openpyxl

PATH = "data/ModRacer.xlsx"
ADD_ROWS = [
    (29, "drift_speed_min", 8.0, "Min speed to enter/hold drift mode (m/s)"),
    (30, "drift_brake_scale", 0.35, "Handbrake force scale while drifting (bite, not full lock)"),
    (31, "drift_rear_grip", 0.62, "Rear lateral grip scale while drifting (front untouched)"),
    (32, "drift_slip_assist", 0.55, "Steering slip assist threshold while drifting (rad)"),
    (33, "drift_yaw_engage", 0.22, "Yaw stability engage angle while drifting (dot domain, ~40deg max drift angle)"),
    (34, "drift_yaw_kick", 0.25, "Drift-entry yaw kick angular velocity (rad/s, x steering input)"),
]

wb = openpyxl.load_workbook(PATH)
ws = wb["Game-game"]
assert ws.cell(row=3, column=1).value == "id" and ws.cell(row=3, column=2).value == "key", \
    "Game 表结构与预期不符，中止"

# 清掉数据区中间/尾部的全空行（历史格式残留会被导出器当成 id=0 空记录）
r = 4
while r <= ws.max_row:
    if all(ws.cell(row=r, column=c).value is None for c in range(1, ws.max_column + 1)):
        ws.delete_rows(r)
    else:
        r += 1

keys = {}
row = 4
while ws.cell(row=row, column=1).value is not None:
    keys[ws.cell(row=row, column=2).value] = row
    row += 1

added = 0
for rid, key, value, note in ADD_ROWS:
    if key in keys:
        print("Game-game 已有 %s，跳过" % key)
        continue
    ws.append([rid, key, value, note])
    added += 1

wb.save(PATH)
print("Game-game 新增 %d 行（drift_*），总行数 %d" % (added, row - 4 + added))
