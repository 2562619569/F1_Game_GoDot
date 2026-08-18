# -*- coding: utf-8 -*-
"""
ModRacer 配表升级脚本（一次性，2026-08）：车-车碰撞冲击放大参数（CollisionKick）。

  - Game-game 补 4 行 bump_* 键（id 14~17）：
      bump_strength   冲量倍率（×碰前接近速度×折合质量）
      bump_min_speed  接近速度死区 m/s（轻蹭不放大）
      bump_max_speed  参与计算的接近速度上限 m/s
      bump_yaw        角落撞击甩尾力矩倍率
保留原有全部数据，按 key 判重幂等。用法:  cd config && python tools/add_bump_tuning.py
"""
import openpyxl

PATH = "data/ModRacer.xlsx"
ROWS = [
    (14, "bump_strength", 0.6, "Car-car impact boost multiplier (x closing speed x reduced mass)"),
    (15, "bump_min_speed", 2.0, "Closing speed below this no boost (m/s)"),
    (16, "bump_max_speed", 25.0, "Closing speed cap for boost math (m/s)"),
    (17, "bump_yaw", 1.0, "Yaw spin torque multiplier on corner hits"),
]

wb = openpyxl.load_workbook(PATH)
ws = wb["Game-game"]
assert ws.cell(row=3, column=1).value == "id" and ws.cell(row=3, column=2).value == "key", \
    "Game 表结构与预期不符，中止"

# 清掉数据区中间/尾部的全空行（历史格式残留会被导出器当成 id=0 空记录）
r = 4
while r <= ws.max_row:
    if all(ws.cell(row=r, column=c).value is None for c in range(1, ws.max_column + 1)):
        ws.delete_rows(r)
    else:
        r += 1

keys = set()
row = 4
while ws.cell(row=row, column=1).value is not None:
    keys.add(ws.cell(row=row, column=2).value)
    row += 1

added = 0
for rid, key, value, note in ROWS:
    if key in keys:
        print("Game-game 已有 %s，跳过" % key)
        continue
    ws.append([rid, key, value, note])
    added += 1

wb.save(PATH)
print("Game-game 新增 %d 行（bump_*），总行数 %d" % (added, row - 4 + added))
