# -*- coding: utf-8 -*-
"""
ModRacer 配表升级脚本（一次性，2026-08）：轮胎改件拆分为「轮毂(外观) + 轮胎(性能+模型)」。

  - 改件-part      补 1 列：model（外观资产目录名，轮胎改件指向 art/tires/<model>/，其余留空）
  - 新建 外观件-cosmetic：纯外观件表（不进背包/装配/掉落，默认全解锁）
      首批 category="wheel" 3 行轮毂（701~ 段），model 指向 art/wheels/<model>/
表结构依据 docs/game-design/配表结构规划.md。保留原有全部数据，可重复执行（幂等）。
用法:  cd config && python tools/add_cosmetic_wheel.py
"""
import openpyxl

PATH = "data/ModRacer.xlsx"

wb = openpyxl.load_workbook(PATH)

# ---------------------------------------------------------------- 改件-part 补 model 列
part = wb["Part-part"]
assert part.cell(row=3, column=18).value == "power", "改件表结构与预期不符，中止"
if part.cell(row=3, column=19).value == "model":
    print("Part-part 已有 model 列，跳过")
else:
    part.cell(row=1, column=19, value="string")
    part.cell(row=2, column=19, value="外观资产目录名(轮胎件指向art/tires/<model>/)")
    part.cell(row=3, column=19, value="model")
    # 轮胎改件 201~204 的对应胎模；其余改件无外观模型，留空串
    TIRE_MODELS = {201: "slick_v1", 202: "offroad_v1", 203: "rain_v1", 204: "allterrain_v1"}
    row = 4
    while part.cell(row=row, column=1).value is not None:
        pid = part.cell(row=row, column=1).value
        part.cell(row=row, column=19, value=TIRE_MODELS.get(pid, ""))
        row += 1
    print("Part-part 补 model 列（轮胎 4 行填值）")

# ---------------------------------------------------------------- 外观件-cosmetic
name = "Cosmetic-cosmetic"
if name in wb.sheetnames:
    wb.remove(wb[name])
ws = wb.create_sheet(name)
ws.append(["int", "tr_string", "string", "int", "string", "tr_string"])
ws.append(["编号", "外观件名", "类别", "稀有度", "资产目录名", "描述"])
ws.append(["id", "name", "category", "rarity", "model", "desc"])
# 701~ 外观件段；category 预留 wheel/paint/spoiler...，首批只有 wheel（轮毂）
# model → art/wheels/<model>/hub.glb（纯外观，无任何属性列）
for r in [
    [701, "Sport Wheels V1", "wheel", 1, "sport_v1", "Box-spoke sport hub. Default look."],
    [702, "Classic Wheels V1", "wheel", 1, "classic_v1", "Classic multi-spoke dish hub."],
    [703, "Aero Wheels V1", "wheel", 1, "aero_v1", "Smooth aero cover hub."],
]:
    ws.append(r)
print("新建 Cosmetic-cosmetic（3 行轮毂）")

wb.save(PATH)
print("已升级", PATH, "：", ", ".join(wb.sheetnames))
