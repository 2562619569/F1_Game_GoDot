# -*- coding: utf-8 -*-
"""
ModRacer 配表升级脚本（2026-08）：车-车碰撞失稳窗口参数（CollisionKick destab_*）。

  - Game-game 补 3 行 destab 键（id 18~20）：
      bump_destab_speed  触发失稳窗口的接近速度阈值 m/s（低于只放大冲量不开窗口）
      bump_destab_time   失稳窗口时长上限 s（随接近速度从 0.4s 线性拉到该值）
      bump_destab_grip   窗口内轮胎摩擦缩放（0.45 ≈ 被撞瞬间轮胎"被震松"）
  - 同步上调既有键：bump_strength 0.6→0.7、bump_yaw 1.0→2.5
    （配合失稳窗口；单独加大冲量只能线性放大晃动，转不成失控）
按 key 判重幂等。用法:  cd config && python tools/add_bump_destab.py
"""
import openpyxl

PATH = "data/ModRacer.xlsx"
ADD_ROWS = [
    (18, "bump_destab_speed", 6.0, "Closing speed to trigger destabilization window (m/s)"),
    (19, "bump_destab_time", 1.0, "Max destabilization window duration (s, scales with closing speed)"),
    (20, "bump_destab_grip", 0.40, "Tire friction scale during destabilization window"),
]
UPDATE_VALUES = {
    "bump_strength": 0.7,
    "bump_yaw": 2.5,
    "bump_destab_grip": 0.40,
}

wb = openpyxl.load_workbook(PATH)
ws = wb["Game-game"]
assert ws.cell(row=3, column=1).value == "id" and ws.cell(row=3, column=2).value == "key", \
    "Game 表结构与预期不符，中止"

# 清掉数据区中间/尾部的全空行（历史格式残留会被导出器当成 id=0 空记录）
r = 4
while r <= ws.max_row:
    if all(ws.cell(row=r, column=c).value is None for c in range(1, ws.max_column + 1)):
        ws.delete_rows(r)
    else:
        r += 1

keys = {}
row = 4
while ws.cell(row=row, column=1).value is not None:
    keys[ws.cell(row=row, column=2).value] = row
    row += 1

added = 0
for rid, key, value, note in ADD_ROWS:
    if key in keys:
        print("Game-game 已有 %s，跳过" % key)
        continue
    ws.append([rid, key, value, note])
    added += 1

updated = 0
for key, value in UPDATE_VALUES.items():
    if key not in keys:
        print("警告: %s 不存在，跳过改值" % key)
        continue
    r = keys[key]
    old = ws.cell(row=r, column=3).value
    if old != value:
        ws.cell(row=r, column=3).value = value
        updated += 1
        print("Game-game %s: %s -> %s" % (key, old, value))

wb.save(PATH)
print("Game-game 新增 %d 行、改值 %d 行（bump_*），总行数 %d" % (added, updated, row - 4 + added))
