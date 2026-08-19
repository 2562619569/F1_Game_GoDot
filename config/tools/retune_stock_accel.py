# -*- coding: utf-8 -*-
"""
ModRacer 配表调整脚本（2026-08）：原厂 0-100 km/h 加速调进 5~7s 区间。

用户反馈无改装加速过慢（实测 601 8.74s / 602 >14s 破不了百 / 603 11.32s，
与车型设定不符）。只动 Car 表 max_torque（物理推力唯一杠杆，齿比/换挡线/
风阻滚阻均不受影响；accel 展示值与改件映射 K_ACCEL_TORQUE 不动）：

  601 Brute Power    420 → 580 Nm   8.74s → 5.56s
  602 Agile Sprinter 260 → 400 Nm   未达  → 6.55s
  603 All-Rounder    330 → 540 Nm   11.32s → 5.19s
  701~703 NPC 交通车同步（npc_check 断言与 601~603 物理逐项一致）

用法:  cd config && python tools/retune_stock_accel.py
"""
import openpyxl

PATH = "data/ModRacer.xlsx"
# car_id → (旧值, 新值)；旧值校验防二次套用
CHANGES = {601: (420.0, 580.0), 602: (260.0, 400.0), 603: (330.0, 540.0),
           701: (420.0, 580.0), 702: (260.0, 400.0), 703: (330.0, 540.0)}

wb = openpyxl.load_workbook(PATH)
ws = wb["Car-car"]
assert ws.cell(row=3, column=1).value == "id", "Car 表结构与预期不符，中止"

# 第 3 行表头定位 max_torque 列
col_torque = None
for c in range(1, ws.max_column + 1):
    if ws.cell(row=3, column=c).value == "max_torque":
        col_torque = c
        break
assert col_torque is not None, "Car 表未找到 max_torque 列，中止"

done = 0
r = 4
while ws.cell(row=r, column=1).value is not None:
    cid = ws.cell(row=r, column=1).value
    if cid in CHANGES:
        old, new = CHANGES[cid]
        cur = float(ws.cell(row=r, column=col_torque).value)
        assert cur == old, "car %d max_torque 当前为 %s（预期 %s），中止防误改" % (cid, cur, old)
        ws.cell(row=r, column=col_torque).value = new
        print("car %d max_torque %s → %s" % (cid, old, new))
        done += 1
    r += 1
assert done == 6, "应改 6 行，实改 %d（表行缺失？）" % done

wb.save(PATH)
print("Car-car 完成 %d 行" % done)
