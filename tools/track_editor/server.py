# ModRacer 赛道编辑器配套服务:静态文件 + 地图 JSON 读写 API
# 一键启动(自动打开浏览器):
#     python tools/track_editor/server.py
# 不打开浏览器(仅起服务):
#     python tools/track_editor/server.py --no-browser
# 双击 file:// 打开编辑器时,API 会请求 http://localhost:8137(本服务须在跑),
# 服务未启动则编辑器自动进入离线模式(仅导入/导出下载)。
import json
import math
import re
import socket
import sys
import threading
import time
import webbrowser
from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler
from pathlib import Path

PORT = 8137
URL = "http://localhost:%d/tools/track_editor/index.html" % PORT
ROOT = Path(__file__).resolve().parent.parent.parent  # 项目根
DATA_DIR = ROOT / "game" / "race" / "tracks" / "data"


def port_responding():
    # Only treat the port as healthy when it actually answers HTTP.
    try:
        import urllib.request
        with urllib.request.urlopen("http://127.0.0.1:%d/api/maps" % PORT, timeout=0.8) as r:
            return r.status == 200
    except Exception:
        return False


def kill_port_owner():
    # Port is occupied but not responding. Usually a stale server.py left a
    # dead listener behind, so terminate only PIDs listening on this port.
    try:
        import os
        import signal
        import subprocess
        out = subprocess.check_output(
            ["netstat", "-ano", "-p", "tcp"],
            stderr=subprocess.STDOUT,
            text=True,
        )
        for line in out.splitlines():
            parts = line.split()
            if len(parts) >= 5 and parts[1].endswith(":%d" % PORT):
                if parts[3] != "LISTENING":
                    continue
                pid = int(parts[-1])
                if pid <= 0:
                    continue
                try:
                    os.kill(pid, signal.SIGTERM)
                except Exception:
                    try:
                        subprocess.run(["taskkill", "/F", "/PID", str(pid)], check=False)
                    except Exception:
                        pass
        time.sleep(0.4)
    except Exception:
        pass



CONFIG_XLSX = ROOT / "config" / "data" / "ModRacer.xlsx"


def read_map_table():
    """Read the Map-map sheet from ModRacer.xlsx as the source of truth."""
    rows_out = []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(CONFIG_XLSX), data_only=True, read_only=True)
        try:
            for ws in wb.worksheets:
                if "map" not in ws.title.lower():
                    continue
                header_idx = {}
                for ri, row in enumerate(ws.iter_rows(values_only=True)):
                    if not header_idx:
                        for ci, cell in enumerate(row):
                            val = str(cell).strip().lower() if cell is not None else ""
                            if val in ("id", "name", "desc"):
                                header_idx[val] = ci
                        if "id" not in header_idx or "name" not in header_idx:
                            continue
                        continue
                    if row[header_idx["id"]] is None:
                        continue
                    get = lambda key: row[header_idx[key]] if header_idx.get(key) is not None and header_idx[key] < len(row) else None
                    rows_out.append({
                        "id": int(get("id")),
                        "name": str(get("name") or ""),
                        "desc": str(get("desc") or ""),
                    })
                break
        finally:
            wb.close()
    except Exception:
        rows_out = []
    return rows_out


def read_env_preset(mid):
    """Map environment preset from map_<id>_env.json (env settings live beside, not inside, map json)."""
    f = DATA_DIR / ("map_%d_env.json" % mid)
    try:
        d = json.loads(f.read_text(encoding="utf-8"))
        return str(d.get("preset", ""))
    except Exception:
        return ""


def default_map_template(meta):
    """Build a minimal editable map JSON from xlsx metadata when no file exists yet."""
    return {
        "version": 1,
        "meta": {"id": meta["id"], "name": meta["name"]},
        "width_default": 24,
        "grid": {"count": 8, "row_gap": 8, "col_gap": 7, "first_row_offset": 6},
        "options": {"walls": True, "wall_height": 0.8, "barrier_offset": 8, "sample_step": 2},
        "routes": [
            {
                "id": "main",
                "surface": "road",
                "points": [
                    {"x": 0, "y": 0, "z": 0, "width": None},
                    {"x": 0, "y": 0, "z": -90, "width": None},
                ],
            }
        ],
        "baked": {},
    }


def baked_route_length(baked):
    """烘焙主线长度(米)。新格式点为 [x,y,z,width] 需累计折线长,
    旧格式 [x,y,z,tx,ty,tz,width,s] 直接取末点弧长 s。"""
    if not baked:
        return 0.0
    last = baked[-1]
    if len(last) >= 8:
        try:
            return float(last[7])
        except (TypeError, ValueError):
            pass
    total = 0.0
    prev = None
    for p in baked:
        cur = (p[0], p[1], p[2])
        if prev is not None:
            total += math.dist(cur, prev)
        prev = cur
    return total


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(ROOT), **kwargs)

    # ---- 基础 ----
    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def _json(self, code, obj):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self._cors()
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(204)
        self._cors()
        self.end_headers()

    # ---- GET: 静态文件 + /api/maps 列表 + /api/maps/<id> 单图 ----
    def do_GET(self):
        if self.path == "/api/maps":
            table = read_map_table()
            files = {}
            if DATA_DIR.is_dir():
                for f in sorted(DATA_DIR.glob("map_*.json")):
                    if f.name.endswith("_env.json"):
                        continue  # 环境文件与地图几何分离，不作为地图列出
                    try:
                        d = json.loads(f.read_text(encoding="utf-8"))
                        baked = d.get("baked", {}).get("main", [])
                        mid = int(d.get("meta", {}).get("id", 0))
                        files[mid] = {
                            "name": str(d.get("meta", {}).get("name", "")),
                            "length": round(baked_route_length(baked)),
                        }
                    except (ValueError, KeyError, IndexError):
                        pass  # skip broken files
            maps = []
            if table:
                for m in table:
                    fid = m["id"]
                    fj = files.get(fid)
                    maps.append({
                        "id": fid,
                        "name": (fj or m)["name"],
                        "weather": read_env_preset(fid),
                        "desc": m.get("desc", ""),
                        "has_file": fid in files,
                        "length": fj["length"] if fj else 0,
                    })
            else:
                for fid, fj in sorted(files.items()):
                    maps.append({
                        "id": fid,
                        "name": fj["name"],
                        "weather": read_env_preset(fid),
                        "desc": "",
                        "has_file": True,
                        "length": fj["length"],
                    })
            return self._json(200, {"maps": maps})
        m = re.match(r"^/api/maps/(\d+)$", self.path)
        if m:
            f = DATA_DIR / ("map_%s.json" % m.group(1))
            if f.is_file():
                try:
                    return self._json(200, json.loads(f.read_text(encoding="utf-8")))
                except ValueError:
                    return self._json(400, {"error": "map json broken"})
            meta = next((row for row in read_map_table() if row["id"] == int(m.group(1))), None)
            if not meta:
                return self._json(404, {"error": "map not found"})
            return self._json(200, default_map_template(meta))
        return super().do_GET()

    # ---- POST: /api/maps/<id> 写回地图 ----
    def do_POST(self):
        m = re.match(r"^/api/maps/(\d+)$", self.path)
        if not m:
            return self._json(404, {"error": "not found"})
        try:
            body = json.loads(self.rfile.read(int(self.headers.get("Content-Length", 0))))
            if not isinstance(body, dict) or "routes" not in body or "baked" not in body:
                raise ValueError("缺少 routes/baked 字段")
        except (ValueError, json.JSONDecodeError) as e:
            return self._json(400, {"error": "bad json: %s" % e})
        # 文件 id 以 URL 为准,防止 meta 与目标文件错位
        body.setdefault("meta", {})["id"] = int(m.group(1))
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        target = DATA_DIR / ("map_%s.json" % m.group(1))
        target.write_text(json.dumps(body, ensure_ascii=False), encoding="utf-8")
        print("[track_editor] saved %s" % target.relative_to(ROOT))
        return self._json(200, {"ok": True, "file": str(target.relative_to(ROOT)).replace("\\", "/")})


if __name__ == "__main__":
    no_browser = "--no-browser" in sys.argv
    if port_responding():
        # Healthy server already running, just open the page.
        if not no_browser:
            webbrowser.open(URL)
        raise SystemExit(0)
    # Port may still be held by a stale listener that does not answer HTTP.
    # On Windows SO_REUSEADDR can allow binding alongside it, so clear it first.
    kill_port_owner()
    try:
        srv = ThreadingHTTPServer(("127.0.0.1", PORT), Handler)
    except OSError:
        print("Port %d is occupied but not responding; clearing stale listener and retrying..." % PORT)
        kill_port_owner()
        srv = ThreadingHTTPServer(("127.0.0.1", PORT), Handler)
    if not no_browser:
        threading.Thread(
            target=lambda: (time.sleep(0.6), webbrowser.open(URL)),
            daemon=True,
        ).start()
    print("ModRacer track editor -> %s  (Ctrl+C 退出)" % URL)
    try:
        srv.serve_forever()
    except KeyboardInterrupt:
        print("\nbye")
