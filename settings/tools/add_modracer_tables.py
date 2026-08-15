# -*- coding: utf-8 -*-
"""
ModRacer 配表升级脚本（一次性，2026-08）。

在既有 data/ModRacer.xlsx（车型/改件/地图 三表）基础上：
  - 车型-car    补 2 列：grip_road / grip_offroad（底盘地形基础能力）
  - 改件-part   补 2 列：effect / power（战术效果枚举与强度）
  - 新建 4 张表：全局-game / 回合-round / 掉落-loot / 排名奖励-rank_reward
表结构依据 docs/game-design/配表结构规划.md。保留原有全部数据，可重复执行（幂等）。
"""
import openpyxl

PATH = "data/ModRacer.xlsx"

wb = openpyxl.load_workbook(PATH)


def sheet_exists(name: str) -> bool:
    return name in wb.sheetnames


def add_columns(ws, start_col: int, types, notes, names):
    """在第 start_col 列起追加三行表头；数据行由调用方逐行填写。"""
    for i, t in enumerate(types):
        ws.cell(row=1, column=start_col + i, value=t)
        ws.cell(row=2, column=start_col + i, value=notes[i])
        ws.cell(row=3, column=start_col + i, value=names[i])


def fill_column(ws, col: int, values_by_row: dict):
    """按 {行号: 值} 写入数据单元格。"""
    for row, v in values_by_row.items():
        ws.cell(row=row, column=col, value=v)


def build_sheet(name: str, types, notes, names, rows):
    if sheet_exists(name):
        wb.remove(wb[name])
    ws = wb.create_sheet(name)
    ws.append(types)
    ws.append(notes)
    ws.append(names)
    for row in rows:
        ws.append(row)


# ---------------------------------------------------------------- 车型-car 补列
car = wb["车型-car"]
assert car.cell(row=3, column=10).value == "desc", "车型表结构与预期不符，中止"
add_columns(car, 11,
            ["int", "int"],
            ["铺装抓地(0-10)", "越野抗性(0-10)"],
            ["grip_road", "grip_offroad"])
fill_column(car, 11, {4: 6, 5: 8, 6: 7})    # RWD 低越野 / FWD 高铺装 / AWD 高越野
fill_column(car, 12, {4: 3, 5: 5, 6: 8})

# ---------------------------------------------------------------- 改件-part 补列
part = wb["改件-part"]
assert part.cell(row=3, column=16).value == "desc", "改件表结构与预期不符，中止"
add_columns(part, 17,
            ["string", "float"],
            ["效果枚举", "效果强度(按effect解释)"],
            ["effect", "power"])
# effect 解释：none=无 / slow_spin=减速打转(强度=角速度deg/s) /
#              stealth=隐身免锁定(强度=半透明度) / nitro_push=氮气推力(强度=推力倍率)
EFFECTS = {
    501: ("slow_spin", 30.0),
    502: ("stealth", 0.5),
    503: ("nitro_push", 2.0),
}
row = 4
while part.cell(row=row, column=1).value is not None:
    pid = part.cell(row=row, column=1).value
    eff, power = EFFECTS.get(pid, ("none", 0.0))
    part.cell(row=row, column=17, value=eff)
    part.cell(row=row, column=18, value=power)
    row += 1

# ---------------------------------------------------------------- 全局-game
build_sheet("全局-game",
            ["int", "string", "float", "string"],
            ["编号", "参数名", "参数值", "说明"],
            ["id", "key", "value", "note"],
            [
                [1, "player_max", 8, "最大玩家数（含 AI）"],
                [2, "intermission_sec", 40, "局间整备时长（秒）"],
                [3, "round_count", 4, "小回合数"],
                [4, "start_countdown", 3, "发车倒计时（秒）"],
                [5, "lock_ahead_range", 60, "火箭筒自动锁定前方距离（米）"],
                [6, "loot_pick_radius", 3, "拾取改件判定半径（米）"],
            ])

# ---------------------------------------------------------------- 回合-round
build_sheet("回合-round",
            ["int", "tr_string", "bool", "int", "string"],
            ["回合号", "回合名", "是否决战局", "时限秒", "候选地图id"],
            ["id", "name", "is_final", "time_limit", "map_pool"],
            [
                [1, "第1回合", False, 240, "1|2"],
                [2, "第2回合", False, 270, "2|3"],
                [3, "第3回合", False, 270, "3|4"],
                [4, "终极决战", True, 300, "1|2|3|4"],
            ])

# ---------------------------------------------------------------- 掉落-loot
build_sheet("掉落-loot",
            ["int", "string", "int", "string", "string", "int"],
            ["编号", "路线类型", "掉落点数", "稀有度权重(r1|r2|r3|r4)", "可掉类别", "保底稀有度(0无)"],
            ["id", "route", "drop_count", "rarity_weights", "category_pool", "guarantee_rarity"],
            [
                [1, "main", 6, "60|30|10|0", "engine|tires|aero|chassis|tactical", 0],
                [2, "hazard", 3, "0|20|50|30", "engine|tires|aero|chassis|tactical", 2],
            ])

# ---------------------------------------------------------------- 排名奖励-rank_reward
build_sheet("排名奖励-rank_reward",
            ["int", "int", "int", "int"],
            ["名次", "奖励件数", "奖励最低稀有度", "下回合发车位"],
            ["id", "reward_count", "reward_rarity_min", "grid_next"],
            [
                [1, 2, 3, 8],
                [2, 2, 2, 7],
                [3, 1, 2, 6],
                [4, 1, 2, 5],
                [5, 1, 1, 4],
                [6, 1, 1, 3],
                [7, 1, 1, 2],
                [8, 1, 1, 1],
            ])

wb.save(PATH)
print("已升级", PATH, "：", ", ".join(wb.sheetnames))
